from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe.utils import flt
from openpyxl import load_workbook

from marina_custom_apps.dc_dispatch.services import history_evidence_service as base
from marina_custom_apps.dc_dispatch.services import history_evidence_light_service as light
from marina_custom_apps.dc_dispatch.services import size_evidence_service


SHEET_NAME = "Score Reconciliation"


def _header_map(sheet):
    return {
        str(cell.value or "").strip(): cell.column
        for cell in sheet[1]
        if cell.value
    }


def _read_base_rows(sheet):
    headers = _header_map(sheet)

    required = [
        "Target Item Template",
        "Final Store",
        "Decision",
        "Reference Store",
        "Raw Demand Score at Final Store",
        "Applied Historical Demand Score",
        "Score Stored on Current Proposal",
    ]

    missing = [
        value
        for value in required
        if value not in headers
    ]
    if missing:
        frappe.throw(
            "Historical Evidence score sheet is missing columns: "
            + ", ".join(missing)
        )

    rows = []
    for row_number in range(2, sheet.max_row + 1):
        target = sheet.cell(
            row_number,
            headers["Target Item Template"],
        ).value
        store = sheet.cell(
            row_number,
            headers["Final Store"],
        ).value

        if not target or not store:
            continue

        rows.append(
            {
                "target": str(target),
                "store": str(store),
                "decision": sheet.cell(
                    row_number,
                    headers["Decision"],
                ).value,
                "reference_store": sheet.cell(
                    row_number,
                    headers["Reference Store"],
                ).value,
                "raw_score": flt(
                    sheet.cell(
                        row_number,
                        headers[
                            "Raw Demand Score at Final Store"
                        ],
                    ).value
                ),
                "reference_adjusted_score": flt(
                    sheet.cell(
                        row_number,
                        headers[
                            "Applied Historical Demand Score"
                        ],
                    ).value
                ),
                "proposal_score": sheet.cell(
                    row_number,
                    headers[
                        "Score Stored on Current Proposal"
                    ],
                ).value,
            }
        )

    return rows


def _expected_scores(run, rows):
    rules = {
        row.store_warehouse: row
        for row in run.store_rules
        if row.decision != "Exclude"
    }

    target_set = {
        row.item_template: (
            str(row.related_set or "").strip()
        )
        for row in run.items
    }

    members_by_set = defaultdict(list)
    for target, related_set in target_set.items():
        if related_set:
            members_by_set[related_set].append(target)

    forecast = {}
    for row in rows:
        rule = rules.get(row["store"])
        growth = (
            flt(getattr(rule, "expected_growth", 0))
            if rule
            else 0
        )
        multiplier = max(
            0.0,
            1.0 + growth / 100.0,
        )
        forecast[
            (row["target"], row["store"])
        ] = max(
            0.0,
            flt(row["reference_adjusted_score"]),
        ) * multiplier

    set_scores = {}

    for related_set, members in members_by_set.items():
        combined = defaultdict(float)

        for target in members:
            member_values = {
                store: forecast.get(
                    (target, store),
                    0.0,
                )
                for store in rules
            }
            total = sum(
                max(0.0, flt(value))
                for value in member_values.values()
            )

            if total <= 0:
                continue

            for store, value in member_values.items():
                combined[store] += (
                    max(0.0, flt(value))
                    / total
                )

        for target in members:
            for store in rules:
                set_scores[
                    (target, store)
                ] = combined.get(
                    store,
                    0.0,
                )

    expected = {}
    for row in rows:
        key = (
            row["target"],
            row["store"],
        )
        related_set = target_set.get(
            row["target"],
            "",
        )

        if related_set:
            score = set_scores.get(
                key,
                0.0,
            )
            basis = (
                "Related Set Combined Score "
                f"({related_set})"
            )
        else:
            score = forecast.get(
                key,
                0.0,
            )
            rule = rules.get(
                row["store"]
            )
            growth = (
                flt(
                    getattr(
                        rule,
                        "expected_growth",
                        0,
                    )
                )
                if rule
                else 0
            )

            if (
                row["decision"]
                == "Use Reference Store"
            ):
                basis = "Reference Store Demand"
            else:
                basis = "Historical Demand"

            if growth:
                basis += (
                    f" + Growth {growth:g}%"
                )

        expected[key] = {
            "score": score,
            "basis": basis,
        }

    return expected, rules


def _rewrite_reconciliation(workbook, run):
    old = workbook[SHEET_NAME]
    rows = _read_base_rows(old)
    expected, rules = _expected_scores(
        run,
        rows,
    )

    index = workbook.sheetnames.index(
        SHEET_NAME
    )
    workbook.remove(old)
    sheet = workbook.create_sheet(
        SHEET_NAME,
        index,
    )

    headers = [
        "Target Item Template",
        "Final Store",
        "Decision",
        "Reference Store",
        "Raw Demand Score at Final Store",
        "Reference-Adjusted Historical Demand",
        "Expected Growth %",
        "Forecast-Adjusted Demand",
        "Proposal Score Basis",
        "Expected Score Used by Proposal",
        "Expected Store Share %",
        "Score Stored on Current Proposal",
        "Reconciles to Proposal",
    ]
    sheet.append(headers)

    expected_totals = defaultdict(float)
    for row in rows:
        key = (
            row["target"],
            row["store"],
        )
        expected_totals[
            row["target"]
        ] += max(
            0.0,
            flt(
                expected.get(
                    key,
                    {},
                ).get(
                    "score",
                    0,
                )
            ),
        )

    for row in rows:
        key = (
            row["target"],
            row["store"],
        )
        rule = rules.get(
            row["store"]
        )
        growth = (
            flt(
                getattr(
                    rule,
                    "expected_growth",
                    0,
                )
            )
            if rule
            else 0
        )
        forecast_adjusted = max(
            0.0,
            row[
                "reference_adjusted_score"
            ],
        ) * max(
            0.0,
            1.0 + growth / 100.0,
        )

        expected_row = expected.get(
            key,
            {
                "score": forecast_adjusted,
                "basis": "Historical Demand",
            },
        )
        expected_score = flt(
            expected_row["score"]
        )
        total = expected_totals[
            row["target"]
        ]
        share = (
            expected_score * 100.0 / total
            if total
            else 0.0
        )

        proposal_score = (
            None
            if row["proposal_score"] is None
            else flt(
                row["proposal_score"]
            )
        )

        reconciles = (
            ""
            if proposal_score is None
            else (
                "Yes"
                if abs(
                    proposal_score
                    - expected_score
                ) < 0.000001
                else "No"
            )
        )

        sheet.append(
            [
                row["target"],
                row["store"],
                row["decision"],
                row["reference_store"],
                row["raw_score"],
                row[
                    "reference_adjusted_score"
                ],
                growth,
                forecast_adjusted,
                expected_row["basis"],
                expected_score,
                share,
                proposal_score,
                reconciles,
            ]
        )

    base._format_sheet(
        sheet,
        freeze="A2",
        auto_filter=True,
    )


@frappe.whitelist()
def download_history_evidence(run_name):
    """Fast standard evidence export for planning validation."""
    run = frappe.get_doc(
        "DC Dispatch Run",
        run_name,
    )
    run.check_permission("read")

    # v0.6.8: use the persisted history cache and omit transaction-heavy
    # Sales by Store / Return Audit sheets from the standard export.
    light.export_lightweight_history_evidence(
        run
    )

    content = (
        frappe.local.response.filecontent
    )
    workbook = load_workbook(
        BytesIO(content),
        data_only=False,
        read_only=False,
    )

    _rewrite_reconciliation(
        workbook,
        run,
    )

    size_evidence_service.add_size_performance_reconciliation(
        workbook,
        run,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filecontent = (
        output.read()
    )
    frappe.local.response.type = "binary"


@frappe.whitelist()
def download_detailed_history_audit(run_name):
    """Optional legacy detailed export containing transaction-heavy sheets."""
    run = frappe.get_doc(
        "DC Dispatch Run",
        run_name,
    )
    run.check_permission("read")

    base.export_history_evidence(
        run
    )

    frappe.local.response.filename = (
        f"{run.name}-Detailed-Historical-Audit"
        + (
            f"-R{run.revision}"
            if run.revision
            else ""
        )
        + ".xlsx"
    )
