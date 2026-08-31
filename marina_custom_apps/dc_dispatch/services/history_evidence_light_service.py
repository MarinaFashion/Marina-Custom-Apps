from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt
from openpyxl import Workbook

from marina_custom_apps.dc_dispatch.services import historical_cache_service as history_cache
from marina_custom_apps.dc_dispatch.services import history_evidence_service as base
from marina_custom_apps.dc_dispatch.services.history_policy_service import (
    historical_filter_fieldnames,
    historical_scope_candidates,
    historical_scope_text,
)
from marina_custom_apps.dc_dispatch.services.run_service import (
    _adjust_store_scores,
    _fields_by_main_group,
    _has_value,
    _item_values,
    _normal,
    _reference_fieldnames,
    assert_calculation_inputs_unchanged,
)


def _cached_demand_sales(run, stores):
    """Use the persisted history cache instead of rescanning Sales Invoices."""
    cache_data = history_cache.load_cache_data(
        run,
        require_valid=True,
    )
    store_set = set(stores or [])

    sales = {}
    for (template, store), quantity in cache_data["demand"].items():
        if store not in store_set:
            continue
        quantity = flt(quantity)
        if quantity != 0:
            sales[(template, store)] = quantity

    return sales


def _match_detail_text(details):
    parts = []
    for detail in details:
        if not detail["comparable"]:
            result = "Target blank / ignored"
        elif detail["matched"]:
            result = "Match"
        else:
            result = "No match"

        parts.append(
            f'{detail["label"]}: target="{detail["target"] or ""}", '
            f'history="{detail["historical"] or ""}" [{result}]'
        )
    return " | ".join(parts)


def _update_summary_notes(sheet):
    replacements = {
        "Cross-Store Returns": (
            "Excluded from demand score. Transaction-level detail is omitted "
            "from the standard evidence export for performance."
        ),
        "Return Resolution": (
            "The proposal uses the validated historical cache built by "
            "Check Store History. Detailed transaction/return audit remains "
            "available through the optional detailed audit endpoint."
        ),
    }

    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row, 1).value
        if key in replacements:
            sheet.cell(row, 2).value = replacements[key]


def export_lightweight_history_evidence(run):
    """Build the standard evidence workbook without transaction-heavy sheets.

    The standard export intentionally omits:
      - Sales by Store
      - Return Audit

    It uses the same persisted historical cache that the proposal calculation
    uses, so exporting evidence does not rescan Sales Invoice / return history.
    """
    if not run.items:
        frappe.throw(
            _("Load target items before exporting historical evidence.")
        )
    if not run.reference_fields:
        frappe.throw(
            _("Select historical matching fields before exporting historical evidence.")
        )
    if not run.store_rules:
        frappe.throw(
            _("Load eligible stores before exporting historical evidence.")
        )

    if run.calculation_input_hash:
        assert_calculation_inputs_unchanged(run)

    settings = frappe.get_single("DC Dispatch Settings")
    active_rules = [
        row
        for row in run.store_rules
        if row.decision != "Exclude"
    ]
    stores = [
        row.store_warehouse
        for row in active_rules
    ]
    if not stores:
        frappe.throw(
            _("No included stores exist in this run.")
        )

    # Performance change: no historical transaction scan on export.
    sales = _cached_demand_sales(
        run,
        stores,
    )
    candidate_templates = {
        template
        for template, _store in sales
    }
    target_templates = {
        row.item_template
        for row in run.items
    }

    value_fields = (
        _reference_fieldnames(run)
        | historical_filter_fieldnames(run)
        | {
            settings.item_main_group_field,
            settings.item_subgroup_field,
            settings.item_related_set_field,
        }
    )
    values = _item_values(
        candidate_templates | target_templates,
        value_fields,
    )
    fields_by_group = _fields_by_main_group(run)
    threshold = flt(run.minimum_match_percent)
    field_labels = base._field_labels(
        value_fields
    )
    proposal_scores = (
        base._current_proposal_scores(run)
    )

    target_summaries = []
    cohort_rows = []
    score_rows = []

    rules_by_store = {
        row.store_warehouse: row
        for row in active_rules
    }

    for item_row in run.items:
        target_values = values.get(
            item_row.item_template,
            {},
        )
        group = (
            target_values.get(
                settings.item_main_group_field
            )
            or item_row.main_group
        )
        selected_fields = (
            fields_by_group.get(
                group,
                [],
            )
        )
        comparable_fields = [
            field
            for field in selected_fields
            if _has_value(
                target_values.get(field)
            )
        ]

        scoped_templates = (
            historical_scope_candidates(
                run,
                item_row,
                candidate_templates,
                values,
            )
        )
        scoped_sales = {
            key: qty
            for key, qty in sales.items()
            if key[0] in scoped_templates
        }

        cohort = []
        match_details = {}

        for candidate in sorted(
            scoped_templates
        ):
            candidate_values = values.get(
                candidate,
                {},
            )
            if (
                candidate_values.get(
                    settings.item_main_group_field
                )
                != group
            ):
                continue

            if not comparable_fields:
                match_percent = 100.0
            else:
                matches = sum(
                    1
                    for field in comparable_fields
                    if (
                        _has_value(
                            candidate_values.get(
                                field
                            )
                        )
                        and _normal(
                            candidate_values.get(
                                field
                            )
                        )
                        == _normal(
                            target_values.get(
                                field
                            )
                        )
                    )
                )
                match_percent = (
                    matches
                    * 100.0
                    / len(comparable_fields)
                )

            if match_percent < threshold:
                continue

            cohort.append(candidate)
            details = []

            for field in selected_fields:
                target_value = (
                    target_values.get(field)
                )
                historical_value = (
                    candidate_values.get(field)
                )
                comparable = _has_value(
                    target_value
                )
                matched = (
                    comparable
                    and _has_value(
                        historical_value
                    )
                    and _normal(
                        historical_value
                    )
                    == _normal(
                        target_value
                    )
                )
                details.append(
                    {
                        "field": field,
                        "label": field_labels.get(
                            field,
                            field,
                        ),
                        "target": target_value,
                        "historical": (
                            historical_value
                        ),
                        "comparable": comparable,
                        "matched": matched,
                    }
                )

            match_details[candidate] = {
                "match_percent": (
                    match_percent
                ),
                "details": details,
            }

        raw_scores = defaultdict(float)
        cohort_set = set(cohort)

        for (
            template,
            store,
        ), quantity in scoped_sales.items():
            if template in cohort_set:
                raw_scores[store] += (
                    quantity
                )

        raw_scores = {
            store: max(
                0,
                quantity,
            )
            for store, quantity
            in raw_scores.items()
        }

        (
            adjusted_scores,
            missing_references,
        ) = _adjust_store_scores(
            run,
            raw_scores,
        )

        adjusted_total = sum(
            max(
                0,
                flt(value),
            )
            for value
            in adjusted_scores.values()
        )

        target_summaries.append(
            {
                "target": (
                    item_row.item_template
                ),
                "main_group": (
                    item_row.main_group
                ),
                "subgroup": (
                    item_row.subgroup
                ),
                "historical_scope": (
                    historical_scope_text(
                        run,
                        item_row.main_group,
                        field_labels,
                    )
                ),
                "scope_templates": len(
                    scoped_templates
                ),
                "selected_fields": ", ".join(
                    field_labels.get(
                        field,
                        field,
                    )
                    for field
                    in selected_fields
                ),
                "comparable_fields": ", ".join(
                    field_labels.get(
                        field,
                        field,
                    )
                    for field
                    in comparable_fields
                ),
                "threshold": threshold,
                "cohort_templates": len(
                    cohort
                ),
                "cohort_units": sum(
                    raw_scores.values()
                ),
                "cohort_stores": sum(
                    1
                    for value
                    in raw_scores.values()
                    if value > 0
                ),
                "missing_references": (
                    ", ".join(
                        missing_references
                    )
                ),
            }
        )

        for candidate in cohort:
            detail = (
                match_details[candidate]
            )
            cohort_rows.append(
                {
                    "target": (
                        item_row.item_template
                    ),
                    "historical": (
                        candidate
                    ),
                    "match_percent": (
                        detail[
                            "match_percent"
                        ]
                    ),
                    "match_detail": (
                        _match_detail_text(
                            detail[
                                "details"
                            ]
                        )
                    ),
                    "historical_demand_units": sum(
                        flt(quantity)
                        for (
                            template,
                            _store,
                        ), quantity
                        in scoped_sales.items()
                        if (
                            template
                            == candidate
                        )
                    ),
                }
            )

        for (
            store,
            rule,
        ) in rules_by_store.items():
            raw = max(
                0,
                flt(
                    raw_scores.get(
                        store,
                        0,
                    )
                ),
            )
            applied = max(
                0,
                flt(
                    adjusted_scores.get(
                        store,
                        0,
                    )
                ),
            )

            score_rows.append(
                {
                    "target": (
                        item_row.item_template
                    ),
                    "store": store,
                    "decision": (
                        rule.decision
                    ),
                    "reference_store": (
                        rule.reference_store
                    ),
                    "raw_score": raw,
                    "applied_score": (
                        applied
                    ),
                    "share_percent": (
                        applied
                        * 100.0
                        / adjusted_total
                        if adjusted_total
                        else 0
                    ),
                    "proposal_score": (
                        proposal_scores.get(
                            (
                                item_row.item_template,
                                store,
                            )
                        )
                    ),
                }
            )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Run Summary"

    base._write_run_summary(
        summary,
        run,
        active_rules,
        field_labels,
    )
    _update_summary_notes(
        summary
    )

    base._write_historical_scope_filters(
        workbook.create_sheet(
            "Historical Scope Filters"
        ),
        run,
        field_labels,
    )
    base._write_target_summary(
        workbook.create_sheet(
            "Target Cohorts"
        ),
        target_summaries,
    )
    base._write_cohort_templates(
        workbook.create_sheet(
            "Historical Templates"
        ),
        cohort_rows,
    )
    base._write_score_reconciliation(
        workbook.create_sheet(
            "Score Reconciliation"
        ),
        score_rows,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filename = (
        f"{run.name}-Historical-Evidence"
        + (
            f"-R{run.revision}"
            if run.revision
            else ""
        )
        + ".xlsx"
    )
    frappe.local.response.filecontent = (
        output.read()
    )
    frappe.local.response.type = "binary"
