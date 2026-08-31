from __future__ import annotations

from collections import defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import cint, flt
from frappe.utils.file_manager import get_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Protection,
)
from openpyxl.utils import get_column_letter

from marina_custom_apps.dc_dispatch.services.allocation import (
    validate_related_sets,
)
from marina_custom_apps.dc_dispatch.services.dispatch_matrix_service import (
    build_dispatch_matrix,
)
from marina_custom_apps.dc_dispatch.services.forecast_service import (
    assert_forecast_configuration_unchanged,
)
from marina_custom_apps.dc_dispatch.services.size_performance_service import (
    assert_size_configuration_unchanged,
)
from marina_custom_apps.dc_dispatch.services.run_service import (
    assert_calculation_inputs_unchanged,
    assert_stock_snapshot,
    validate_current_proposal,
)


SIMPLE_SHEET = "Simple Allocation"
DATA_SHEET = "_Data"

VISIBLE_FIXED_HEADERS = [
    "Item Template",
    "Size",
]

TOTAL_HEADERS = [
    "Total DC Qty",
    "Total Dispatched",
    "Remaining Qty",
]

TECHNICAL_HEADERS = [
    "__Item Code",
    "__Run ID",
    "__Revision",
]


@frappe.whitelist()
def download_proposal(run_name):
    run = frappe.get_doc(
        "DC Dispatch Run",
        run_name,
    )
    run.check_permission("read")
    return export_proposal(run)


def export_proposal(run):
    if run.status not in {
        "Calculated",
        "Proposal Imported",
    }:
        frappe.throw(
            _("Calculate the proposal before exporting it.")
        )

    assert_calculation_inputs_unchanged(run)
    assert_size_configuration_unchanged(run)
    assert_forecast_configuration_unchanged(run)

    lines = _proposal_lines(run)
    if not lines:
        frappe.throw(
            _("No proposal lines exist for this revision.")
        )

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Run Summary"
    _write_summary(
        summary,
        run,
        lines,
    )

    _write_style_summary(
        workbook.create_sheet("Style Summary"),
        run,
    )

    matrix = build_dispatch_matrix(
        run,
        lines=lines,
    )

    data_sheet = workbook.create_sheet(DATA_SHEET)
    _write_hidden_data(
        data_sheet,
        run,
        matrix,
    )

    _write_simple_allocation(
        workbook.create_sheet(SIMPLE_SHEET),
        run,
        matrix,
    )

    _write_warnings(
        workbook.create_sheet("Warnings"),
        run,
    )

    # Technical data must remain in the workbook for safe re-import,
    # but normal users do not need to see it.
    data_sheet.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filename = (
        f"{run.name}-R{run.revision}-DC-Dispatch.xlsx"
    )
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"


def import_proposal(run):
    """Import the editable Simple Allocation matrix as Final Qty.

    User-facing rules:
      - only store quantity cells are intended to be edited;
      - Tier Min/Max are proposal-generation rules and may be overridden;
      - original Dispatch % / target may also be overridden by the reviewed
        matrix;
      - the hard stock ceiling remains the DC Stock Snapshot per variant.
    """
    if run.status not in {
        "Calculated",
        "Proposal Imported",
    }:
        frappe.throw(
            _("Only a calculated proposal can be imported.")
        )

    if not run.proposal_file:
        frappe.throw(
            _("Attach the reviewed proposal workbook first.")
        )

    assert_calculation_inputs_unchanged(run)
    assert_size_configuration_unchanged(run)
    assert_forecast_configuration_unchanged(run)
    assert_stock_snapshot(run)

    _filename, content = get_file(
        run.proposal_file
    )

    try:
        workbook = load_workbook(
            BytesIO(content),
            data_only=False,
            read_only=False,
        )
    except Exception as exc:
        frappe.throw(
            _(
                "The attached file is not a valid Excel "
                "workbook: {0}"
            ).format(exc)
        )

    if SIMPLE_SHEET not in workbook.sheetnames:
        frappe.throw(
            _(
                "The workbook does not contain the "
                "Simple Allocation sheet."
            )
        )

    sheet = workbook[SIMPLE_SHEET]

    database_lines = frappe.get_all(
        "DC Dispatch Proposal Line",
        filters={
            "run": run.name,
            "revision": run.revision,
        },
        fields=[
            "name",
            "store_warehouse",
            "item_template",
            "item_code",
            "related_set",
            "suggested_qty",
            "final_qty",
            "exclude",
        ],
        limit_page_length=0,
    )

    if not database_lines:
        frappe.throw(
            _("No proposal lines exist for the current revision.")
        )

    matrix = build_dispatch_matrix(
        run,
        lines=database_lines,
    )

    expected_stores = list(
        matrix["stores"]
    )
    expected_rows = {
        row["item_code"]: row
        for row in matrix["rows"]
    }

    headers = {
        cell.value: cell.column
        for cell in sheet[1]
        if cell.value
    }

    required_headers = (
        VISIBLE_FIXED_HEADERS
        + expected_stores
        + TOTAL_HEADERS
        + TECHNICAL_HEADERS
    )

    missing_headers = [
        header
        for header in required_headers
        if header not in headers
    ]

    if missing_headers:
        frappe.throw(
            _(
                "The Simple Allocation sheet is missing columns: {0}"
            ).format(
                ", ".join(missing_headers)
            )
        )

    # Store columns must remain exactly the current active stores.
    # v0.6.5 layout keeps A:E fixed, so stores start immediately after
    # Remaining Qty and end immediately before the hidden technical columns.
    first_store_col = headers["Remaining Qty"] + 1
    last_store_col = headers["__Item Code"] - 1
    actual_store_headers = [
        sheet.cell(1, column).value
        for column in range(
            first_store_col,
            last_store_col + 1,
        )
    ]

    if actual_store_headers != expected_stores:
        frappe.throw(
            _(
                "Store columns were changed. Please export a fresh "
                "workbook and edit only the store quantity cells."
            )
        )

    line_by_key = {
        (
            row.item_code,
            row.store_warehouse,
        ): row
        for row in database_lines
    }

    imported = {}
    seen_item_codes = set()
    variant_totals = defaultdict(int)
    template_totals = defaultdict(int)
    validation_rows = []
    errors = []

    for row_number in range(
        2,
        sheet.max_row + 1,
    ):
        item_code = sheet.cell(
            row_number,
            headers["__Item Code"],
        ).value

        # Ignore the visible bottom Total row / blank rows.
        if not item_code:
            continue

        item_code = str(item_code).strip()

        if item_code in seen_item_codes:
            errors.append(
                f"Row {row_number}: duplicate variant {item_code}."
            )
            continue

        seen_item_codes.add(item_code)

        if item_code not in expected_rows:
            errors.append(
                f"Row {row_number}: unknown variant {item_code}."
            )
            continue

        run_id = str(
            sheet.cell(
                row_number,
                headers["__Run ID"],
            ).value
            or ""
        ).strip()

        revision = cint(
            sheet.cell(
                row_number,
                headers["__Revision"],
            ).value
        )

        if (
            run_id != run.name
            or revision != cint(run.revision)
        ):
            errors.append(
                f"Row {row_number}: Run ID or Revision does not "
                "match the current proposal."
            )
            continue

        expected = expected_rows[item_code]

        visible_template = str(
            sheet.cell(
                row_number,
                headers["Item Template"],
            ).value
            or ""
        ).strip()

        visible_size = str(
            sheet.cell(
                row_number,
                headers["Size"],
            ).value
            or ""
        ).strip()

        if visible_template != str(
            expected["item_template"] or ""
        ).strip():
            errors.append(
                f"Row {row_number}: Item Template was changed."
            )
            continue

        if visible_size != str(
            expected["size"] or ""
        ).strip():
            errors.append(
                f"Row {row_number}: Size was changed."
            )
            continue

        row_total = 0

        for store in expected_stores:
            value = sheet.cell(
                row_number,
                headers[store],
            ).value

            try:
                number = float(
                    0
                    if value in (None, "")
                    else value
                )
            except (TypeError, ValueError):
                errors.append(
                    f"Row {row_number}, {store}: quantity must "
                    "be a whole number."
                )
                continue

            if (
                number < 0
                or not number.is_integer()
            ):
                errors.append(
                    f"Row {row_number}, {store}: quantity must "
                    "be a non-negative whole number."
                )
                continue

            quantity = cint(number)
            key = (
                item_code,
                store,
            )

            original = line_by_key.get(key)
            if not original:
                errors.append(
                    f"Row {row_number}, {store}: no proposal line "
                    f"exists for {item_code}."
                )
                continue

            imported[original.name] = {
                "final_qty": quantity,
                # Zero Final Qty is sufficient; the old Exclude column
                # is no longer part of the user workflow.
                "exclude": 0,
                "override_reason": "",
            }

            row_total += quantity
            variant_totals[item_code] += quantity
            template_totals[
                original.item_template
            ] += quantity

            validation_rows.append(
                {
                    "store_warehouse": (
                        original.store_warehouse
                    ),
                    "item_template": (
                        original.item_template
                    ),
                    "related_set": (
                        original.related_set
                    ),
                    "final_qty": quantity,
                    "exclude": 0,
                }
            )

    missing_variants = (
        set(expected_rows)
        - seen_item_codes
    )

    if missing_variants:
        errors.append(
            "The workbook is missing variants: "
            + ", ".join(
                sorted(missing_variants)[:30]
            )
        )

    if len(imported) != len(database_lines):
        errors.append(
            "The workbook does not contain exactly one editable "
            "quantity for every proposal variant/store line."
        )

    if errors:
        frappe.throw(
            "<br>".join(errors[:50])
        )

    snapshots = {
        row.item_code: cint(
            flt(row.actual_qty)
        )
        for row in frappe.get_all(
            "DC Dispatch Stock Snapshot",
            filters={
                "run": run.name,
                "revision": run.revision,
            },
            fields=[
                "item_code",
                "actual_qty",
            ],
            limit_page_length=0,
        )
    }

    stock_errors = []

    for item_code, quantity in (
        variant_totals.items()
    ):
        available = snapshots.get(
            item_code,
            0,
        )
        if quantity > available:
            stock_errors.append(
                f"{item_code}: Final Qty {quantity} exceeds "
                f"DC Stock Snapshot {available}."
            )

    if stock_errors:
        frappe.throw(
            "<br>".join(stock_errors[:50])
        )

    # Preserve the Related Set integrity rule. Tier Min/Max are
    # intentionally NOT validated here because manual review is
    # allowed to override them.
    expected_members = defaultdict(set)
    for row in run.items:
        if row.related_set:
            expected_members[
                row.related_set
            ].add(row.item_template)

    related_errors = validate_related_sets(
        validation_rows,
        expected_members,
    )

    if related_errors:
        frappe.throw(
            "<br>".join(related_errors[:50])
        )

    # Promote the reviewed style total to the run's current Dispatch Target.
    # This makes Dispatch % an initial planning input while still allowing
    # the reviewed matrix to become the final approved plan.
    for item_row in run.items:
        item_row.target_qty = cint(
            template_totals.get(
                item_row.item_template,
                0,
            )
        )

    for line_id, values in imported.items():
        frappe.db.set_value(
            "DC Dispatch Proposal Line",
            line_id,
            values,
            update_modified=False,
        )

    run.status = "Proposal Imported"
    run.save()

    # Normal proposal validation now succeeds because target_qty has been
    # promoted to the reviewed final style total. Variant stock remains the
    # hard ceiling.
    validate_current_proposal(run)

    return {
        "lines": len(imported),
        "variants": len(expected_rows),
        "stores": len(expected_stores),
        "final_qty": sum(
            variant_totals.values()
        ),
        "status": run.status,
    }


def _proposal_lines(run):
    return frappe.get_all(
        "DC Dispatch Proposal Line",
        filters={
            "run": run.name,
            "revision": run.revision,
        },
        fields=[
            "name",
            "run",
            "revision",
            "source_warehouse",
            "store_warehouse",
            "transit_warehouse",
            "item_template",
            "item_code",
            "related_set",
            "main_group",
            "sales_score",
            "share_percent",
            "suggested_qty",
            "final_qty",
            "exclude",
            "override_reason",
            "validation_status",
        ],
        order_by=(
            "item_template asc, "
            "store_warehouse asc, "
            "item_code asc"
        ),
        limit_page_length=0,
    )


def _write_summary(
    sheet,
    run,
    lines,
):
    current_final = sum(
        0
        if row.exclude
        else cint(row.final_qty)
        for row in lines
    )

    rows = [
        ("DC Dispatch Run", run.name),
        ("Revision", run.revision),
        ("Company", run.company),
        ("Source DC", run.source_warehouse),
        (
            "Sales Period",
            (
                f"{run.sales_from_date} "
                f"to {run.sales_to_date}"
            ),
        ),
        ("Status at Export", run.status),
        ("Styles", len(run.items)),
        (
            "Eligible Stores",
            sum(
                1
                for row in run.store_rules
                if row.decision != "Exclude"
            ),
        ),
        (
            "Current Final Quantity",
            current_final,
        ),
        (
            "Size Performance Factor",
            (
                "Yes"
                if cint(
                    getattr(
                        run,
                        "include_size_performance_factor",
                        0,
                    )
                )
                else "No"
            ),
        ),
        (
            "Size Performance Weight %",
            (
                flt(
                    getattr(
                        run,
                        "size_performance_weight",
                        0,
                    )
                )
                if cint(
                    getattr(
                        run,
                        "include_size_performance_factor",
                        0,
                    )
                )
                else 0
            ),
        ),
    ]

    for row in rows:
        sheet.append(row)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 48

    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _write_style_summary(
    sheet,
    run,
):
    headers = [
        "Item Template",
        "Main Group",
        "Subgroup",
        "Related Set",
        "Available DC Qty",
        "Dispatch %",
        "Dispatch Target",
        "Matching Templates",
        "Cohort Net Units",
        "Stores with Sales",
        "Warning",
    ]

    sheet.append(headers)

    for row in run.items:
        sheet.append(
            [
                row.item_template,
                row.main_group,
                row.subgroup,
                row.related_set,
                row.dc_qty,
                row.dispatch_percentage,
                row.target_qty,
                row.cohort_templates,
                row.cohort_units,
                row.cohort_stores,
                row.warning,
            ]
        )

    _format_table(
        sheet,
        editable_columns=set(),
    )


def _write_hidden_data(
    sheet,
    run,
    matrix,
):
    sheet.append(
        [
            "Item Code",
            "Item Template",
            "Size",
            "Snapshot DC Qty",
            "Run ID",
            "Revision",
        ]
    )

    for row in matrix["rows"]:
        sheet.append(
            [
                row["item_code"],
                row["item_template"],
                row["size"],
                row["total_dc_qty"],
                run.name,
                cint(run.revision),
            ]
        )

    # The technical sheet is not protected because it is hidden.
    # Import never trusts its formulas or totals; it validates against DB.
    for column in range(
        1,
        sheet.max_column + 1,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 20


def _write_simple_allocation(
    sheet,
    run,
    matrix,
):
    stores = list(
        matrix["stores"]
    )

    # Keep the five planner reference columns together at the left:
    # Item Template, Size, Total Dispatched, Total DC Qty, Remaining Qty.
    # Store columns start after them so Excel can freeze A:E as one block.
    headers = (
        VISIBLE_FIXED_HEADERS
        + TOTAL_HEADERS
        + stores
        + TECHNICAL_HEADERS
    )
    sheet.append(headers)

    item_code_col = len(headers) - 2
    run_id_col = len(headers) - 1
    revision_col = len(headers)

    total_dc_col = 3
    total_dispatched_col = 4
    remaining_col = 5

    first_store_col = 6
    last_store_col = (
        first_store_col
        + len(stores)
        - 1
    )

    for row_index, row in enumerate(
        matrix["rows"],
        start=2,
    ):
        sheet.cell(
            row_index,
            1,
            row["item_template"],
        )
        sheet.cell(
            row_index,
            2,
            row["size"],
        )

        for offset, store in enumerate(
            stores,
            start=first_store_col,
        ):
            sheet.cell(
                row_index,
                offset,
                cint(
                    row["store_quantities"].get(
                        store,
                        0,
                    )
                ),
            )

        first_store_letter = (
            get_column_letter(
                first_store_col
            )
        )
        last_store_letter = (
            get_column_letter(
                last_store_col
            )
        )

        sheet.cell(
            row_index,
            total_dispatched_col,
            (
                f"=SUM("
                f"{first_store_letter}{row_index}:"
                f"{last_store_letter}{row_index})"
            ),
        )

        item_code_letter = (
            get_column_letter(
                item_code_col
            )
        )

        # Formula intentionally points to the hidden snapshot sheet.
        # Server import independently validates against the DB snapshot.
        sheet.cell(
            row_index,
            total_dc_col,
            (
                f"=SUMIF("
                f"'{DATA_SHEET}'!$A:$A,"
                f"${item_code_letter}{row_index},"
                f"'{DATA_SHEET}'!$D:$D)"
            ),
        )

        total_dispatched_letter = (
            get_column_letter(
                total_dispatched_col
            )
        )
        total_dc_letter = (
            get_column_letter(
                total_dc_col
            )
        )

        sheet.cell(
            row_index,
            remaining_col,
            (
                f"={total_dc_letter}{row_index}-"
                f"{total_dispatched_letter}{row_index}"
            ),
        )

        sheet.cell(
            row_index,
            item_code_col,
            row["item_code"],
        )
        sheet.cell(
            row_index,
            run_id_col,
            run.name,
        )
        sheet.cell(
            row_index,
            revision_col,
            cint(run.revision),
        )

    total_row = sheet.max_row + 1

    sheet.cell(
        total_row,
        1,
        "Total",
    )

    # Store totals
    for column in range(
        first_store_col,
        last_store_col + 1,
    ):
        letter = get_column_letter(
            column
        )
        sheet.cell(
            total_row,
            column,
            (
                f"=SUM("
                f"{letter}2:"
                f"{letter}{total_row - 1})"
            ),
        )

    # Total Dispatched, Total DC Qty and Remaining are also formulas.
    for column in (
        total_dispatched_col,
        total_dc_col,
        remaining_col,
    ):
        letter = get_column_letter(
            column
        )
        sheet.cell(
            total_row,
            column,
            (
                f"=SUM("
                f"{letter}2:"
                f"{letter}{total_row - 1})"
            ),
        )

    _format_simple_allocation(
        sheet,
        stores,
        first_store_col,
        last_store_col,
        total_dispatched_col,
        total_dc_col,
        remaining_col,
        item_code_col,
        run_id_col,
        revision_col,
        total_row,
    )


def _format_simple_allocation(
    sheet,
    stores,
    first_store_col,
    last_store_col,
    total_dispatched_col,
    total_dc_col,
    remaining_col,
    item_code_col,
    run_id_col,
    revision_col,
    total_row,
):
    header_fill = PatternFill(
        "solid",
        fgColor="1F7A35",
    )
    editable_fill = PatternFill(
        "solid",
        fgColor="C6EFCE",
    )
    alternate_editable_fill = PatternFill(
        "solid",
        fgColor="A9E6B2",
    )
    locked_fill = PatternFill(
        "solid",
        fgColor="E2F0D9",
    )
    warning_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )
    total_fill = PatternFill(
        "solid",
        fgColor="1F7A35",
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.protection = Protection(
            locked=True
        )

    for row_number in range(
        2,
        total_row,
    ):
        editable_row_fill = (
            editable_fill
            if row_number % 2 == 0
            else alternate_editable_fill
        )

        # Item Template / Size are locked reference values.
        for column in (1, 2):
            cell = sheet.cell(
                row_number,
                column,
            )
            cell.fill = locked_fill
            cell.protection = Protection(
                locked=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Only store quantity cells are editable.
        for column in range(
            first_store_col,
            last_store_col + 1,
        ):
            cell = sheet.cell(
                row_number,
                column,
            )
            cell.fill = editable_row_fill
            cell.protection = Protection(
                locked=False
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.number_format = "0"

        # Formula totals are locked.
        for column in (
            total_dispatched_col,
            total_dc_col,
            remaining_col,
        ):
            cell = sheet.cell(
                row_number,
                column,
            )
            cell.fill = (
                warning_fill
                if column == remaining_col
                else locked_fill
            )
            cell.protection = Protection(
                locked=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.number_format = "0"

        # Hidden technical values are locked.
        for column in (
            item_code_col,
            run_id_col,
            revision_col,
        ):
            sheet.cell(
                row_number,
                column,
            ).protection = Protection(
                locked=True
            )

    for cell in sheet[total_row]:
        cell.fill = total_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.protection = Protection(
            locked=True
        )

    sheet.freeze_panes = "F2"

    if total_row > 2:
        sheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(remaining_col)}"
            f"{total_row - 1}"
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 10

    for column in range(
        first_store_col,
        last_store_col + 1,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 18

    for column in (
        total_dispatched_col,
        total_dc_col,
        remaining_col,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = 16

    # Technical columns remain in the workbook but are invisible.
    for column in (
        item_code_col,
        run_id_col,
        revision_col,
    ):
        sheet.column_dimensions[
            get_column_letter(column)
        ].hidden = True

    # Worksheet protection intentionally disabled for reliable editing across Excel clients.


def _write_warnings(
    sheet,
    run,
):
    sheet.append(
        [
            "Type",
            "Record",
            "Warning / Decision",
        ]
    )

    for row in run.items:
        if row.warning:
            sheet.append(
                [
                    "Weak Historical Evidence",
                    row.item_template,
                    row.warning,
                ]
            )

    for row in run.store_rules:
        if row.history_status == "No History":
            decision = row.decision

            if row.reference_store:
                decision += (
                    f": {row.reference_store}"
                )

            sheet.append(
                [
                    "Store Without History",
                    row.store_warehouse,
                    decision,
                ]
            )

    _format_table(
        sheet,
        editable_columns=set(),
    )


def _format_table(
    sheet,
    editable_columns,
):
    header_fill = PatternFill(
        "solid",
        fgColor="551C25",
    )
    editable_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )

    header_map = {
        cell.value: cell.column
        for cell in sheet[1]
    }

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for header, column in (
        header_map.items()
    ):
        width = min(
            45,
            max(
                12,
                len(str(header)) + 2,
            ),
        )

        for cell in sheet[
            get_column_letter(column)
        ]:
            if cell.row > 1:
                width = min(
                    45,
                    max(
                        width,
                        len(
                            str(
                                cell.value or ""
                            )
                        )
                        + 2,
                    ),
                )

                if header in editable_columns:
                    cell.protection = Protection(
                        locked=False
                    )
                    cell.fill = editable_fill
                else:
                    cell.protection = Protection(
                        locked=True
                    )

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = width
