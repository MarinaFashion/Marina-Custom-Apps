"""Sell-through focused complete-range allocation for Stock Allocation Run."""

from __future__ import annotations

from collections import Counter
from io import BytesIO
from math import floor

import frappe
from frappe import _
from frappe.utils import add_days, cint, flt, nowdate, now_datetime
from frappe.utils.file_manager import save_file

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from marina_custom_apps.stock_auto_allocation.sell_through_logic import (
    daily_velocity,
    days_cover,
    donor_is_commercially_valid,
    protected_qty,
    required_qty,
)
from marina_custom_apps.stock_auto_allocation.doctype.stock_allocation_run.stock_allocation_run import (
    StockAllocationRun as BaseStockAllocationRun,
    _get_effective_stock,
    _get_store_warehouses,
    _get_transit_warehouse,
    _sum_sales_qty,
)


class FullScopeStockAllocationRun(BaseStockAllocationRun):
    """Reallocate stock to maximize sell-through without breaking productive stores."""

    def validate(self):
        super().validate()

        if cint(self.lookback_period_days) < 1:
            self.lookback_period_days = 14
        if cint(self.coverage_days) < 1:
            self.coverage_days = 7
        if cint(self.source_protection_days) < 0:
            self.source_protection_days = 3
        if cint(self.default_minimum_per_variant) < 1:
            self.default_minimum_per_variant = 1

        for row in self.items or []:
            if cint(row.minimum_per_variant) < 0:
                row.minimum_per_variant = 0

        scope = self.allocation_scope or "Many-to-Many"

        if scope in ("One-to-Many", "One-to-One") and not self.selected_source_store:
            frappe.throw(_("Select the Source Store for {0}.").format(scope))

        if scope in ("Many-to-One", "One-to-One") and not self.selected_target_store:
            frappe.throw(_("Select the Target Store for {0}.").format(scope))

        if (
            self.selected_source_store
            and self.selected_target_store
            and self.selected_source_store == self.selected_target_store
        ):
            frappe.throw(_("Source and Target stores must be different."))

    @frappe.whitelist()
    def generate_proposal(self):
        if not self.items:
            frappe.throw(_("Pull items into the working list first (Get Items)."))
        if not self.dc_warehouse:
            frappe.throw(_("Set the DC Warehouse before generating a proposal."))

        stores = _get_store_warehouses(self.company)
        if not stores:
            frappe.throw(
                _('No warehouses are flagged "Is Store" for company {0}.').format(
                    self.company
                )
            )

        scope = self.allocation_scope or "Many-to-Many"

        if scope in ("One-to-Many", "One-to-One"):
            if self.selected_source_store not in stores:
                frappe.throw(_("The selected source warehouse is not a store."))

        if scope in ("Many-to-One", "One-to-One"):
            if self.selected_target_store not in stores:
                frappe.throw(_("The selected target warehouse is not a store."))

        self.proposal_lines = []
        self.proposal_version = cint(self.proposal_version) + 1
        self.proposal_review_status = "Not Reviewed"
        self.proposal_exported_on = None
        self.proposal_reviewed_by = None
        self.proposal_reviewed_on = None
        lookback_start = add_days(nowdate(), -cint(self.lookback_period_days))
        default_minimum = max(1, cint(self.default_minimum_per_variant))
        skipped = []
        unfulfilled = []

        for row in self.items:
            if row.excluded:
                continue

            variants = frappe.get_all(
                "Item",
                filters={"variant_of": row.item_template, "disabled": 0},
                pluck="name",
                order_by="name asc",
            )
            if not variants:
                continue

            item_minimum = cint(row.minimum_per_variant) or default_minimum
            missing, reason = self._plan_style(
                template=row.item_template,
                variants=variants,
                stores=stores,
                minimum_per_variant=item_minimum,
                lookback_start=lookback_start,
                scope=scope,
            )
            unfulfilled.extend(missing)
            if reason:
                skipped.append((row.item_template, reason))

        self.status = "Proposal Generated"
        self.save()

        self._show_summary(skipped, unfulfilled)

    def _show_summary(self, skipped, unfulfilled):
        if skipped:
            counts = Counter(reason for _, reason in skipped)
            labels = {
                "No Target Sales": _("no target sales"),
                "Insufficient Range Supply": _("insufficient stock to complete the minimum range"),
                "No Transfer Required": _("no transfer required"),
            }
            parts = [
                f"{count} {labels.get(reason, reason)}"
                for reason, count in counts.items()
            ]
            frappe.msgprint(
                _("{0} style(s) were skipped: {1}.").format(
                    len(skipped),
                    "; ".join(parts),
                ),
                indicator="orange",
                alert=True,
            )

        if unfulfilled:
            preview = [
                f"{item} → {store} short {qty}"
                for item, store, qty in unfulfilled[:12]
            ]
            more = (
                f" (+{len(unfulfilled) - 12} more)"
                if len(unfulfilled) > 12
                else ""
            )
            frappe.msgprint(
                _("Some commercially justified needs could not be covered: {0}{1}").format(
                    "; ".join(preview),
                    more,
                ),
                indicator="orange",
                alert=True,
            )

    def _plan_style(
        self,
        template,
        variants,
        stores,
        minimum_per_variant,
        lookback_start,
        scope,
    ):
        source_stores, target_stores, use_dc = self._scope_stores(scope, stores)

        lookback_days = max(1, cint(self.lookback_period_days))
        target_coverage_days = max(1, cint(self.coverage_days))
        source_protection_days = max(0, cint(self.source_protection_days))

        source_stock = {}
        target_stock = {}
        sales = {}
        style_sales = {}

        relevant = list(dict.fromkeys(source_stores + target_stores))

        for store in relevant:
            style_total = 0.0
            for variant in variants:
                source_stock[(store, variant)] = self._effective_qty(
                    variant,
                    store,
                    bool(self.consider_transit_at_source),
                )
                target_stock[(store, variant)] = self._effective_qty(
                    variant,
                    store,
                    bool(self.consider_transit_at_target),
                )
                sold = max(
                    0.0,
                    flt(
                        _sum_sales_qty(
                            [variant],
                            warehouse=store,
                            from_date=lookback_start,
                        )
                    ),
                )
                sales[(store, variant)] = sold
                style_total += sold

            style_sales[store] = style_total

        eligible_targets = [
            store for store in target_stores if style_sales.get(store, 0) > 0
        ]
        if not eligible_targets:
            return [], "No Target Sales"

        source_protected = {}
        target_required = {}

        for source in source_stores:
            for variant in variants:
                source_protected[(source, variant)] = protected_qty(
                    sales_qty=sales[(source, variant)],
                    lookback_days=lookback_days,
                    protection_days=source_protection_days,
                    minimum_per_variant=minimum_per_variant,
                )

        for target in eligible_targets:
            for variant in variants:
                target_required[(target, variant)] = required_qty(
                    sales_qty=sales[(target, variant)],
                    lookback_days=lookback_days,
                    coverage_days=target_coverage_days,
                    minimum_per_variant=minimum_per_variant,
                )

        dc_initial = {
            variant: (
                self._effective_qty(
                    variant,
                    self.dc_warehouse,
                    bool(self.consider_transit_at_source),
                )
                if use_dc
                else 0
            )
            for variant in variants
        }

        selected_targets = self._select_targets_for_complete_range(
            scope=scope,
            eligible_targets=eligible_targets,
            variants=variants,
            minimum_per_variant=minimum_per_variant,
            source_stores=source_stores,
            source_stock=source_stock,
            source_protected=source_protected,
            target_stock=target_stock,
            style_sales=style_sales,
            dc_available=dc_initial,
        )

        if not selected_targets:
            return [], "Insufficient Range Supply"

        selected_targets = sorted(
            selected_targets,
            key=lambda store: (-style_sales.get(store, 0), store),
        )

        sim_source = dict(source_stock)
        sim_target = dict(target_stock)
        dc_remaining = dict(dc_initial)
        proposal_count_before = len(self.proposal_lines)
        unfulfilled = []

        # Phase 1: restore the minimum complete range everywhere possible.
        for target in selected_targets:
            for variant in variants:
                need = max(
                    0,
                    minimum_per_variant - sim_target[(target, variant)],
                )
                if need <= 0:
                    continue

                remaining = self._allocate_need(
                    template=template,
                    variant=variant,
                    target=target,
                    needed=need,
                    range_phase=True,
                    selected_targets=selected_targets,
                    source_stores=source_stores,
                    source_stock=source_stock,
                    original_target_stock=target_stock,
                    sim_source=sim_source,
                    sim_target=sim_target,
                    sales=sales,
                    style_sales=style_sales,
                    source_protected=source_protected,
                    target_required=target_required,
                    dc_initial=dc_initial,
                    dc_remaining=dc_remaining,
                    use_dc=use_dc,
                    minimum_per_variant=minimum_per_variant,
                    lookback_days=lookback_days,
                    target_coverage_days=target_coverage_days,
                )
                if remaining > 0:
                    unfulfilled.append((variant, target, remaining))

        # Phase 2: add depth according to sales velocity and target coverage.
        depth_needs = []
        for target in selected_targets:
            for variant in variants:
                need = max(
                    0,
                    target_required[(target, variant)]
                    - sim_target[(target, variant)],
                )
                if need <= 0:
                    continue
                velocity = daily_velocity(
                    sales[(target, variant)],
                    lookback_days,
                )
                current_cover = days_cover(
                    sim_target[(target, variant)],
                    sales[(target, variant)],
                    lookback_days,
                )
                depth_needs.append(
                    (
                        target,
                        variant,
                        need,
                        velocity,
                        current_cover if current_cover is not None else 10**12,
                    )
                )

        depth_needs.sort(
            key=lambda row: (
                row[4],       # lowest current cover first
                -row[3],      # then highest velocity
                -style_sales.get(row[0], 0),
                row[0],
                row[1],
            )
        )

        for target, variant, need, _, _ in depth_needs:
            remaining = self._allocate_need(
                template=template,
                variant=variant,
                target=target,
                needed=need,
                range_phase=False,
                selected_targets=selected_targets,
                source_stores=source_stores,
                source_stock=source_stock,
                original_target_stock=target_stock,
                sim_source=sim_source,
                sim_target=sim_target,
                sales=sales,
                style_sales=style_sales,
                source_protected=source_protected,
                target_required=target_required,
                dc_initial=dc_initial,
                dc_remaining=dc_remaining,
                use_dc=use_dc,
                minimum_per_variant=minimum_per_variant,
                lookback_days=lookback_days,
                target_coverage_days=target_coverage_days,
            )
            if remaining > 0:
                unfulfilled.append((variant, target, remaining))

        if (
            len(self.proposal_lines) == proposal_count_before
            and not unfulfilled
        ):
            return [], "No Transfer Required"

        return unfulfilled, None

    def _scope_stores(self, scope, stores):
        source_stores = list(stores)
        target_stores = list(stores)
        use_dc = True

        if scope == "One-to-Many":
            source_stores = [self.selected_source_store]
            target_stores = [
                store for store in stores
                if store != self.selected_source_store
            ]
            use_dc = False

        elif scope == "Many-to-One":
            source_stores = [
                store for store in stores
                if store != self.selected_target_store
            ]
            target_stores = [self.selected_target_store]
            use_dc = True

        elif scope == "One-to-One":
            source_stores = [self.selected_source_store]
            target_stores = [self.selected_target_store]
            use_dc = False

        return source_stores, target_stores, use_dc

    def _select_targets_for_complete_range(
        self,
        scope,
        eligible_targets,
        variants,
        minimum_per_variant,
        source_stores,
        source_stock,
        source_protected,
        target_stock,
        style_sales,
        dc_available,
    ):
        # Explicit target modes always keep the requested target. The proposal
        # may still report an uncovered size if supply is insufficient.
        if scope in ("Many-to-One", "One-to-One"):
            return list(eligible_targets)

        ranked = sorted(
            eligible_targets,
            key=lambda store: (-style_sales.get(store, 0), store),
        )
        selected = []

        supply = {
            variant: dc_available.get(variant, 0)
            + sum(
                max(
                    0,
                    source_stock[(source, variant)]
                    - source_protected[(source, variant)],
                )
                for source in source_stores
            )
            for variant in variants
        }

        cumulative_need = {variant: 0 for variant in variants}

        for candidate in ranked:
            candidate_need = {
                variant: max(
                    0,
                    minimum_per_variant
                    - target_stock[(candidate, variant)],
                )
                for variant in variants
            }

            feasible = all(
                cumulative_need[variant] + candidate_need[variant]
                <= supply[variant]
                for variant in variants
            )

            # Stores already carrying a complete minimum range cost no transfer
            # capacity, so they remain selected even when external supply is tight.
            already_complete = all(
                candidate_need[variant] == 0 for variant in variants
            )

            if feasible or already_complete:
                selected.append(candidate)
                for variant in variants:
                    cumulative_need[variant] += candidate_need[variant]

        return selected

    def _allocate_need(
        self,
        template,
        variant,
        target,
        needed,
        range_phase,
        selected_targets,
        source_stores,
        source_stock,
        original_target_stock,
        sim_source,
        sim_target,
        sales,
        style_sales,
        source_protected,
        target_required,
        dc_initial,
        dc_remaining,
        use_dc,
        minimum_per_variant,
        lookback_days,
        target_coverage_days,
    ):
        remaining = max(0, int(needed))

        if remaining <= 0:
            return 0

        # DC is the first source only in Many-to-Many and Many-to-One.
        if use_dc and dc_remaining.get(variant, 0) > 0:
            qty = min(remaining, dc_remaining[variant])
            if qty > 0:
                self._append_proposal(
                    template=template,
                    variant=variant,
                    source=self.dc_warehouse,
                    destination=target,
                    qty=qty,
                    tier="DC",
                    source_stock=dc_initial.get(variant, 0),
                    source_sales=0,
                    source_protection_qty=0,
                    target_stock=original_target_stock[(target, variant)],
                    target_sales=sales[(target, variant)],
                    target_required_qty=target_required[(target, variant)],
                    source_days_cover_value=None,
                    target_days_cover_value=days_cover(
                        original_target_stock[(target, variant)],
                        sales[(target, variant)],
                        lookback_days,
                    ),
                )
                dc_remaining[variant] -= qty
                sim_target[(target, variant)] += qty
                remaining -= qty

        if remaining <= 0:
            return 0

        donors = []
        for source in source_stores:
            if source == target:
                continue

            floor_qty = source_protected[(source, variant)]

            # In the depth phase, a store that is also a selected receiver keeps
            # its own target requirement before donating further depth.
            if not range_phase and source in selected_targets:
                floor_qty = max(
                    floor_qty,
                    target_required.get((source, variant), minimum_per_variant),
                )

            available = max(
                0,
                sim_source[(source, variant)] - floor_qty,
            )
            if available <= 0:
                continue

            commercially_valid = donor_is_commercially_valid(
                source_variant_sales=sales[(source, variant)],
                target_variant_sales=sales[(target, variant)],
                source_style_sales=style_sales.get(source, 0),
                target_style_sales=style_sales.get(target, 0),
                source_stock=sim_source[(source, variant)],
                lookback_days=lookback_days,
                target_coverage_days=target_coverage_days,
                range_phase=range_phase,
            )
            if not commercially_valid:
                continue

            cover = days_cover(
                sim_source[(source, variant)],
                sales[(source, variant)],
                lookback_days,
            )
            donors.append(
                (
                    source,
                    available,
                    cover,
                    floor_qty,
                )
            )

        donors.sort(
            key=lambda row: (
                0 if sales[(row[0], variant)] <= 0 else 1,
                sales[(row[0], variant)],
                style_sales.get(row[0], 0),
                -(row[2] if row[2] is not None else 10**12),
                _distance_between(row[0], target),
                row[0],
            )
        )

        for source, available, cover, floor_qty in donors:
            if remaining <= 0:
                break

            qty = min(remaining, available)
            if qty <= 0:
                continue

            target_cover_before = days_cover(
                original_target_stock[(target, variant)],
                sales[(target, variant)],
                lookback_days,
            )

            self._append_proposal(
                template=template,
                variant=variant,
                source=source,
                destination=target,
                qty=qty,
                tier="Store",
                source_stock=source_stock[(source, variant)],
                source_sales=sales[(source, variant)],
                source_protection_qty=floor_qty,
                target_stock=original_target_stock[(target, variant)],
                target_sales=sales[(target, variant)],
                target_required_qty=target_required[(target, variant)],
                source_days_cover_value=cover,
                target_days_cover_value=target_cover_before,
            )

            sim_source[(source, variant)] -= qty
            # If the source is also a receiver, its effective target stock must
            # reflect stock it has committed to send away.
            if (source, variant) in sim_target:
                sim_target[(source, variant)] = max(
                    0,
                    sim_target[(source, variant)] - qty,
                )

            sim_target[(target, variant)] += qty
            remaining -= qty

        return remaining

    @frappe.whitelist()
    def export_proposal_for_review(self):
        """Create a controlled XLSX copy of the current proposal for manager review."""
        if self.status != "Proposal Generated":
            frappe.throw(_("Export is available only after a proposal has been generated."))
        if not self.proposal_lines:
            frappe.throw(_("There are no proposal lines to export."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Proposal Review"

        headers = [
            "Proposal Line ID",
            "Allocation Run",
            "Proposal Version",
            "Item Template",
            "Item Variant",
            "Source Warehouse",
            "Target Warehouse",
            "Proposed Qty",
            "Reviewed Qty",
            "Review Action",
            "Review Comment",
            "Tier",
            "Source Stock",
            "Source Sales",
            "Source Protected Qty",
            "Source Days Cover",
            "Target Stock",
            "Target Sales",
            "Target Required Qty",
            "Target Days Cover",
        ]
        ws.append(headers)

        locked_fill = PatternFill("solid", fgColor="E7E6E6")
        editable_fill = PatternFill("solid", fgColor="FFF2CC")
        header_fill = PatternFill("solid", fgColor="551C25")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)

        for line in self.proposal_lines:
            ws.append(
                [
                    line.name,
                    self.name,
                    cint(self.proposal_version),
                    line.item_template,
                    line.item_code,
                    line.source_warehouse,
                    line.target_warehouse,
                    flt(line.qty),
                    flt(line.qty),
                    "Approve",
                    "",
                    line.tier,
                    flt(line.source_stock),
                    flt(line.source_sales),
                    flt(line.source_protection_qty),
                    line.source_days_cover,
                    flt(line.target_stock),
                    flt(line.target_sales),
                    flt(line.target_required_qty),
                    line.target_days_cover,
                ]
            )

        editable_columns = {9, 10, 11}  # Reviewed Qty, Action, Comment
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row, start=1):
                if idx in editable_columns:
                    cell.fill = editable_fill
                    cell.protection = Protection(locked=False)
                else:
                    cell.fill = locked_fill
                    cell.protection = Protection(locked=True)

        action_validation = DataValidation(
            type="list",
            formula1='"Approve,Adjust,Reject"',
            allow_blank=False,
        )
        ws.add_data_validation(action_validation)
        action_validation.add(f"J2:J{ws.max_row}")

        widths = {
            "A": 20, "B": 18, "C": 16, "D": 18, "E": 20, "F": 28, "G": 28,
            "H": 14, "I": 14, "J": 16, "K": 38, "L": 10, "M": 14, "N": 14,
            "O": 20, "P": 18, "Q": 14, "R": 14, "S": 20, "T": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.protection.sheet = True
        ws.protection.enable()

        instructions = wb.create_sheet("Instructions")
        instructions["A1"] = "Stock Allocation Proposal Review"
        instructions["A1"].font = Font(bold=True, size=14, color="551C25")
        instructions["A3"] = "Editable columns"
        instructions["A4"] = "Reviewed Qty"
        instructions["B4"] = "Final quantity requested by the Inventory Manager."
        instructions["A5"] = "Review Action"
        instructions["B5"] = "Approve, Adjust, or Reject."
        instructions["A6"] = "Review Comment"
        instructions["B6"] = "Required for Adjust and Reject."
        instructions["A8"] = "Rules"
        rules = [
            "Approve: Reviewed Qty must equal Proposed Qty.",
            "Adjust: Reviewed Qty must be between 0 and Proposed Qty and a comment is required.",
            "Reject: Reviewed Qty must be 0 and a comment is required.",
            "Do not add, remove, reorder identifiers, or change system-calculated columns.",
            "The file can be uploaded only to the same Allocation Run and Proposal Version.",
            "If the proposal is regenerated after export, the old Excel file becomes invalid.",
        ]
        for i, rule in enumerate(rules, start=9):
            instructions[f"A{i}"] = f"• {rule}"
        instructions.column_dimensions["A"].width = 90
        instructions.column_dimensions["B"].width = 70

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{self.name}-proposal-v{cint(self.proposal_version)}-review.xlsx"
        file_doc = save_file(
            filename,
            output.getvalue(),
            self.doctype,
            self.name,
            is_private=1,
        )

        self.proposal_review_status = "Exported"
        self.proposal_exported_on = now_datetime()
        self.save()

        return {
            "file_url": file_doc.file_url,
            "file_name": filename,
            "proposal_version": cint(self.proposal_version),
        }

    @frappe.whitelist()
    def import_reviewed_proposal(self, file_url):
        """Validate and import manager decisions from a reviewed proposal XLSX."""
        if self.status != "Proposal Generated":
            frappe.throw(_("Reviewed proposals can be uploaded only while the run is in Proposal Generated status."))
        if not file_url:
            frappe.throw(_("Upload an XLSX file first."))

        files = frappe.get_all(
            "File",
            filters={
                "file_url": file_url,
                "attached_to_doctype": self.doctype,
                "attached_to_name": self.name,
            },
            pluck="name",
            limit=1,
        )
        if not files:
            frappe.throw(_("The uploaded file must be attached to this Stock Allocation Run."))

        content = frappe.get_doc("File", files[0]).get_content()
        try:
            wb = load_workbook(BytesIO(content), data_only=True)
        except Exception:
            frappe.throw(_("The uploaded file is not a valid XLSX workbook."))

        if "Proposal Review" not in wb.sheetnames:
            frappe.throw(_('The workbook must contain a sheet named "Proposal Review".'))

        ws = wb["Proposal Review"]
        header_map = {
            str(cell.value).strip(): idx
            for idx, cell in enumerate(ws[1], start=1)
            if cell.value is not None
        }
        required_headers = {
            "Proposal Line ID",
            "Allocation Run",
            "Proposal Version",
            "Proposed Qty",
            "Reviewed Qty",
            "Review Action",
            "Review Comment",
        }
        missing_headers = sorted(required_headers - set(header_map))
        if missing_headers:
            frappe.throw(_("Missing required review columns: {0}").format(", ".join(missing_headers)))

        current_lines = {line.name: line for line in self.proposal_lines}
        imported = {}
        errors = []

        for row_no in range(2, ws.max_row + 1):
            line_id = ws.cell(row_no, header_map["Proposal Line ID"]).value
            if not line_id:
                continue
            line_id = str(line_id).strip()

            if line_id in imported:
                errors.append(f"Row {row_no}: duplicate Proposal Line ID {line_id}.")
                continue
            if line_id not in current_lines:
                errors.append(f"Row {row_no}: Proposal Line ID {line_id} does not belong to the current proposal.")
                continue

            run_name = str(ws.cell(row_no, header_map["Allocation Run"]).value or "").strip()
            version = cint(ws.cell(row_no, header_map["Proposal Version"]).value)
            proposed_in_file = flt(ws.cell(row_no, header_map["Proposed Qty"]).value)
            reviewed_qty = flt(ws.cell(row_no, header_map["Reviewed Qty"]).value)
            action = str(ws.cell(row_no, header_map["Review Action"]).value or "").strip().title()
            comment = str(ws.cell(row_no, header_map["Review Comment"]).value or "").strip()
            line = current_lines[line_id]

            if run_name != self.name:
                errors.append(f"Row {row_no}: Allocation Run does not match {self.name}.")
            if version != cint(self.proposal_version):
                errors.append(
                    f"Row {row_no}: Proposal Version {version} is obsolete; current version is {cint(self.proposal_version)}."
                )
            if abs(proposed_in_file - flt(line.qty)) > 0.0001:
                errors.append(f"Row {row_no}: Proposed Qty was changed in Excel.")
            if reviewed_qty < 0 or reviewed_qty - flt(line.qty) > 0.0001:
                errors.append(
                    f"Row {row_no}: Reviewed Qty must be between 0 and Proposed Qty ({flt(line.qty)})."
                )
            if action not in {"Approve", "Adjust", "Reject"}:
                errors.append(f"Row {row_no}: Review Action must be Approve, Adjust, or Reject.")
            elif action == "Approve" and abs(reviewed_qty - flt(line.qty)) > 0.0001:
                errors.append(f"Row {row_no}: Approve requires Reviewed Qty to equal Proposed Qty.")
            elif action == "Reject" and reviewed_qty != 0:
                errors.append(f"Row {row_no}: Reject requires Reviewed Qty = 0.")
            if action in {"Adjust", "Reject"} and not comment:
                errors.append(f"Row {row_no}: Review Comment is required for {action}.")

            imported[line_id] = {
                "reviewed_qty": reviewed_qty,
                "action": action,
                "comment": comment,
            }

        missing_lines = sorted(set(current_lines) - set(imported))
        if missing_lines:
            errors.append(
                "The workbook is missing {0} proposal line(s). Do not delete proposal rows from the review file.".format(
                    len(missing_lines)
                )
            )

        if errors:
            preview = "<br>".join(f"• {frappe.utils.escape_html(error)}" for error in errors[:25])
            more = (
                f"<br>• ... and {len(errors) - 25} more error(s)."
                if len(errors) > 25
                else ""
            )
            frappe.throw(_("The reviewed proposal could not be imported:<br>{0}{1}").format(preview, more))

        reviewed_on = now_datetime()
        for line_id, values in imported.items():
            line = current_lines[line_id]
            line.reviewed_qty = values["reviewed_qty"]
            line.review_action = values["action"]
            line.review_comment = values["comment"]
            line.reviewed_by = frappe.session.user
            line.reviewed_on = reviewed_on

        self.proposal_review_status = "Reviewed"
        self.proposal_reviewed_by = frappe.session.user
        self.proposal_reviewed_on = reviewed_on
        self.save()

        approved = sum(1 for values in imported.values() if values["reviewed_qty"] > 0)
        rejected = len(imported) - approved
        adjusted = sum(1 for values in imported.values() if values["action"] == "Adjust")
        return {
            "approved_or_adjusted_lines": approved,
            "adjusted_lines": adjusted,
            "rejected_lines": rejected,
        }

    @frappe.whitelist()
    def approve(self):
        if self.status != "Proposal Generated":
            frappe.throw(_('Only a run with status "Proposal Generated" can be approved.'))
        if not self.proposal_lines:
            frappe.throw(_("There are no proposal lines to approve."))

        reviewed = self.proposal_review_status == "Reviewed"
        approved_count = 0

        for line in self.proposal_lines:
            if not line.transit_warehouse:
                frappe.throw(
                    _("Line for {0} -> {1} has no Transit Warehouse resolved.").format(
                        line.item_code,
                        line.target_warehouse,
                    )
                )

            if reviewed:
                final_qty = flt(line.reviewed_qty)
                if line.review_action == "Reject" or final_qty <= 0:
                    line.status = "Rejected"
                    continue
            line.status = "Approved"
            approved_count += 1

        if not approved_count:
            frappe.throw(_("The proposal contains no quantity approved for transfer."))

        self.status = "Approved"
        self.save()

    @frappe.whitelist()
    def create_material_requests(self):
        if self.status != "Approved":
            frappe.throw(_("Only an approved run can have Material Requests created."))

        reviewed = self.proposal_review_status == "Reviewed"
        groups = {}
        for line in self.proposal_lines:
            if line.status != "Approved":
                continue

            final_qty = flt(line.reviewed_qty) if reviewed else flt(line.qty)
            if final_qty <= 0:
                continue

            key = (line.source_warehouse, line.transit_warehouse)
            groups.setdefault(key, []).append((line, final_qty))

        if not groups:
            frappe.throw(_("There are no approved quantities to create Material Requests for."))

        errors = []
        created = 0
        for (source, transit), rows in groups.items():
            try:
                mr = frappe.new_doc("Material Request")
                mr.material_request_type = "Material Transfer"
                mr.company = self.company
                mr.schedule_date = nowdate()
                mr.stock_auto_allocation_run = self.name

                for line, final_qty in rows:
                    mr.append(
                        "items",
                        {
                            "item_code": line.item_code,
                            "qty": final_qty,
                            "warehouse": transit,
                            "from_warehouse": source,
                            "schedule_date": nowdate(),
                        },
                    )

                mr.insert(ignore_permissions=True)
                mr.submit()
                created += 1

                for line, _ in rows:
                    line.status = "Requested"
                    line.material_request = mr.name
            except Exception:
                frappe.log_error(
                    frappe.get_traceback(),
                    f"Stock Allocation Run {self.name}: MR creation failed",
                )
                errors.append(f"{source} -> {transit}")

        if created:
            self.status = "Requested"
        self.save()

        if errors:
            frappe.msgprint(
                _("Some Material Requests could not be created and were skipped: {0}. Check the Error Log.").format(
                    ", ".join(errors)
                ),
                indicator="orange",
                alert=True,
            )

    def on_trash(self):
        """Clear generated-document backlinks before Frappe validates links.

        Frappe calls on_trash before check_if_doc_is_linked(), so this avoids
        the circular delete dependency without changing Link field types or
        globally disabling link protection.
        """
        mr_names = {
            line.material_request
            for line in self.proposal_lines
            if line.material_request
        }
        for mr_name in mr_names:
            if frappe.db.exists("Material Request", mr_name):
                frappe.db.set_value(
                    "Material Request",
                    mr_name,
                    "stock_auto_allocation_run",
                    "",
                    update_modified=False,
                )

        # Defensive cleanup for the future Transit -> Store leg.
        if frappe.get_meta("Stock Entry").has_field("stock_auto_allocation_run"):
            frappe.db.set_value(
                "Stock Entry",
                {"stock_auto_allocation_run": self.name},
                "stock_auto_allocation_run",
                "",
                update_modified=False,
            )

    @staticmethod
    def _effective_qty(item_code, warehouse, consider_transit):
        return max(
            0,
            floor(
                flt(
                    _get_effective_stock(
                        item_code,
                        warehouse,
                        consider_transit,
                    )
                )
            ),
        )

    def _append_proposal(
        self,
        template,
        variant,
        source,
        destination,
        qty,
        tier,
        source_stock,
        source_sales,
        source_protection_qty,
        target_stock,
        target_sales,
        target_required_qty,
        source_days_cover_value,
        target_days_cover_value,
    ):
        if qty <= 0 or source == destination:
            return

        self.append(
            "proposal_lines",
            {
                "item_template": template,
                "item_code": variant,
                "source_warehouse": source,
                "target_warehouse": destination,
                "transit_warehouse": _get_transit_warehouse(destination),
                "qty": qty,
                "tier": tier,
                "status": "Proposed",
                "source_stock": source_stock,
                "source_sales": source_sales,
                "source_protection_qty": source_protection_qty,
                "source_days_cover": source_days_cover_value,
                "target_stock": target_stock,
                "target_sales": target_sales,
                "target_required_qty": target_required_qty,
                "target_days_cover": target_days_cover_value,
            },
        )


def _distance_between(source, destination):
    if source == destination:
        return 0

    rows = frappe.get_all(
        "Store Distance",
        filters=[
            ["from_store", "in", [source, destination]],
            ["to_store", "in", [source, destination]],
        ],
        fields=["from_store", "to_store", "distance_km"],
    )
    for row in rows:
        if {row.from_store, row.to_store} == {source, destination}:
            return flt(row.distance_km)

    return 10**12
